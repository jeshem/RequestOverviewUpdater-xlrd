import xlrd
from xlutils.copy import copy
import openpyxl

class Read_Write_File(object):

    def __init__(self, loc, ov_file):
        self.location = loc
        self.overviewfile = ov_file

        self.file_list=[]
        self.project_keys = []
        self.service_keys = []

        self.project_data = {}
        self.service_data = {}

    #append extracted information to the request overview fil
    @staticmethod
    def get_keys_from_init(self, location, proj_keys, serv_keys):
        section_name = ""

        #the config file, "/nsoci.ini", should be in the same directory as the request forms and overview file
        #change here if you want to rename the config file and/or change whereit is located
        with open(location + "/nsoci.ini") as fp:
            line = fp.readline()
            while line:
                s = line.strip()
           
                # skip blank line
                if not s == "":          
                    if section_name == "Projects":

                        if not '[' in s:
                            # extract keys under [Projects] and then continue
                            proj_keys.append(s)                    

                    elif section_name == "Services":
                        if not '[' in s:
                            # extract keys under [Services] and then continue
                            serv_keys.append(s)

                    if '[' in s:
                        section_name = self.check_section(s)
           
                line = fp.readline()
        return proj_keys, serv_keys

    @staticmethod
    def check_section(name):
        print ("\ncheck for section: " + name) 
    
        if name == "[Projects]":
            return "Projects"
    
        elif name == "[Services]":
            return "Services"

    
    #using the keys found in the config file, pull data from the request forms
    @staticmethod
    def read_from_excel(file, proj_keys, serv_keys):
        print ("reading from " + str(file))

        rb = xlrd.open_workbook(file)
        sheet = rb.sheet_by_index(0)

        proj_data = {}
        serv_data = {}        

        error = False

        #get data from the request forms
        for search in proj_keys:
            key_row = -1
            key_col = -1

            for r in range(sheet.nrows):
                for c in range(sheet.ncols):
                    cell = sheet.cell(r, c)
                    if cell.value == search + ":":
                        key_row = r
                        key_col = c
                        break

            #If the request form being searched does not contain the Project keys,
            #print an error message and exit the loop
            if key_row == -1 or key_col == -1:
                   print ("ERROR: " + file + " does not contain the key " + search)
                   error = True
                   break
            
            data = sheet.cell(key_row, key_col + 6).value

            #Make sure project information in the request form are filled out
            if data == None:
                    print("ERROR: " + file + " is missing Project information missing for " + search)
                    error = True
                    break
            else:
                proj_data[search] = data
    
        for search in serv_keys:
            key_row = -1
            key_col = -1

            if error:
                break

            for r in range(sheet.nrows):
                for c in range(sheet.ncols):
                    cell = sheet.cell(r, c)
                    if str(cell.value).strip() == search:
                        key_row = r
                        key_col = c
                        break

            #If the request form being searched does not contain the Service keys,
            #print an error message and exit the loop
            if key_row == -1 or key_col == -1:
                   print ("ERROR: " + file + " does not contain the key " + search)
                   error = True
                   break

            service_values = sheet.cell(key_row, key_col + 7)

            if service_values.value == "Not to be requested":
                data = None
            else:
                data = service_values.value
            serv_data[search] = data

        #If request form is incomplete or does not contain the required keys, return empty dictionaries
        if error:
            proj_data = {}
            serv_data = {}

        return proj_data, serv_data

    #Using the data pulled from the request form, write to the overview file
    @staticmethod
    def write_to_excel(loc, overviewfile, proj_keys, serv_keys, proj_data, serv_data):
        wb = openpyxl.load_workbook(loc + "\\" + overviewfile + ".xlsx")
        ws = wb["OCI Resources"]

        #change these if the rows in the overview file start are shifted or if the number of rows exceed 500
        current_row = 6
        last_row = 0

        is_last_row = False

        '''
        Note that the overview file must have an empty row at the end of the list of project names in order
        for the program to find where to insert the new row. If there is a row in the middle of the file
        missing its project name, the program will insert the new project before that row.
        '''
        while not is_last_row:
            if ws.cell(current_row, 1).value == None:
                ws.insert_rows(current_row)
                is_last_row = True
            else:
                current_row +=1
    
        #find the columns that the keys are in and insert the corresponding data into those columns
        for i in range(len(proj_keys)):
            print(str(i))
            '''
            Some of the keys taken from the request form do not match the column titles so I had to hardcode it
            to look for the correct titles upon coming across those keys.
            Alternatively, the request form and overview file can be updated to have matching keys/column titles.
            '''
            if proj_keys[i] == "Project requestor":
                search = "Resource requestor"
            else:
                search = proj_keys[i]
                
            for col_index in range(1, ws.max_column):
                if ws.cell(row=5, column = col_index).value == search:
                    key_col = col_index
                    col_index = ws.max_column
                print(ws.cell(row=5, column = col_index).value)

            ws.cell(row=current_row, column=key_col).value = proj_data[proj_keys[i]]
        
        for i in range (len(serv_keys)):
            print(str(i))

            for col_index in range(1, ws.max_column):
                if ws.cell(row=5, column = col_index).value == serv_keys[i]:
                    key_col = col_index
                    col_index = ws.max_column
                col_index += 1

            ws.cell(row=current_row, column=key_col).value = serv_data[serv_keys[i]]

        wb.save(loc + "\\" + overviewfile + ".xlsx")

    def read_write(self, file_list):
        self.project_keys, self.service_keys = self.get_keys_from_init(self, self.location, self.project_keys, self.service_keys)
        
        for file in file_list:
            self.project_data, self.service_data = self.read_from_excel(file, self.project_keys, self.service_keys)
            if self.project_data and self.service_data:
                self.write_to_excel(self.location, self.overviewfile, self.project_keys, self.service_keys, self.project_data, self.service_data)


